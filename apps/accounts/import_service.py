import re

import openpyxl
from unidecode import unidecode

from apps.classes.models import (
    Enrollment,
)
from .services import (
    StudentReferenceService,
)
from .models import User


class StudentImportService:

    @staticmethod
    def generate_username(
        full_name,
    ):

        username = (
            unidecode(full_name)
            .lower()
            .strip()
        )

        username = re.sub(
            r"[^a-z0-9]",
            "",
            username,
        )

        base_username = username

        counter = 1

        while User.objects.filter(
            username=username
        ).exists():

            counter += 1

            username = (
                f"{base_username}{counter}"
            )

        return username
    @staticmethod
    def import_students(
        *,
        excel_file,
        school_class,
        default_password="123456",
    ):

        workbook = openpyxl.load_workbook(
            excel_file
        )

        sheet = workbook.active

        created_count = 0
        skipped_count = 0

        issues = []

        for row in range(
            3,
            sheet.max_row + 1,
        ):

            raw_name = (
                sheet.cell(
                    row=row,
                    column=2,
                ).value
            )

            # EMPTY ROW
            if raw_name is None:

                issues.append(
                    f"Row {row}: empty name"
                )

                skipped_count += 1
                continue

            full_name = (
                str(raw_name)
                .replace("\xa0", " ")
                .strip()
                .title()
            )

            # INVALID SHORT NAME
            if len(full_name) < 2:

                issues.append(
                    (
                        f"Row {row}: "
                        f"invalid short name "
                        f"'{full_name}'"
                    )
                )

                skipped_count += 1
                continue

            # NUMERIC NAME
            if full_name.isnumeric():

                issues.append(
                    (
                        f"Row {row}: "
                        f"numeric value "
                        f"'{full_name}'"
                    )
                )

                skipped_count += 1
                continue
            existing_student = (
                User.objects.filter(
                    role=User.Role.STUDENT,
                    full_name__iexact=full_name,
                ).first()
            )
            # DUPLICATE
            if existing_student:

                issues.append(
                    (
                        f"Row {row}: "
                        f"duplicate student -> "
                        f"Excel='{full_name}' | "
                        f"DB='{existing_student.full_name}' | "
                        f"username='{existing_student.username}' | "
                        f"reference='{existing_student.reference_number}'"
                    )
                )

                print(
                    "DUPLICATE:",
                    {
                        "row": row,
                        "excel_name": full_name,
                        "db_name": existing_student.full_name,
                        "username": existing_student.username,
                        "reference": existing_student.reference_number,
                    }
                )

                skipped_count += 1
                continue
            
            username = (
                StudentImportService.generate_username(
                    full_name
                )
            )
            next_number = (
                StudentReferenceService.get_next_number()
            )

            reference_number = (
                StudentReferenceService.generate_reference_number(
                    next_number
                )
            )

            student = User.objects.create_user(
                username=username,
                password=default_password,
                full_name=full_name,
                role=User.Role.STUDENT,
                reference_number=reference_number,
            )

            Enrollment.objects.create(
                student=student,
                school_class=school_class,
            )

            created_count += 1

        return {
            "created_count": created_count,
            "skipped_count": skipped_count,
            "issues": issues,
        }