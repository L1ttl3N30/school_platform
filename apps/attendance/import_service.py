from calendar import monthrange
from datetime import date

import openpyxl

from apps.accounts.models import User
from apps.core.utils import (
    normalize_vietnamese_text,
)

from .models import (
    AttendanceRecord,
    AttendanceStatus,
)


class AttendanceImportService:

    START_DAY_COLUMN = 3
    END_DAY_COLUMN = 33

    @classmethod
    def find_student(
        cls,
        student_name,
    ):

        normalized_excel_name = (
            normalize_vietnamese_text(
                student_name
            )
        )

        students = User.objects.filter(
            role=User.Role.STUDENT
        )

        for student in students:

            normalized_db_name = (
                normalize_vietnamese_text(
                    student.full_name
                )
            )

            if (
                normalized_db_name
                == normalized_excel_name
            ):
                return student

        return None

    @classmethod
    def import_excel(
        cls,
        *,
        excel_file,
        school_class,
        month,
        year,
        marked_by,
    ):

        workbook = openpyxl.load_workbook(
            excel_file
        )

        sheet = workbook.active

        imported_count = 0
        skipped_count = 0

        unmatched_students = []

        max_days = monthrange(
            year,
            month,
        )[1]

        for row in range(
            3,
            sheet.max_row + 1,
        ):

            student_name = (
                sheet.cell(
                    row=row,
                    column=2,
                ).value
            )

            if not student_name:

                skipped_count += 1

                unmatched_students.append(
                    f"Row {row}: empty name"
                )

                continue

            student_name = str(
                student_name
            ).strip()

            student = cls.find_student(
                student_name
            )

            if not student:

                skipped_count += 1

                unmatched_students.append(
                    (
                        f"Row {row}: "
                        f"student not found -> "
                        f"{student_name}"
                    )
                )

                continue

            for column in range(
                cls.START_DAY_COLUMN,
                cls.END_DAY_COLUMN + 1,
            ):

                day = (
                    column
                    - cls.START_DAY_COLUMN
                    + 1
                )

                if day > max_days:
                    continue

                cell_value = (
                    sheet.cell(
                        row=row,
                        column=column,
                    ).value
                )

                if (
                    str(cell_value).strip()
                    != "1"
                ):
                    continue

                attendance_date = date(
                    year,
                    month,
                    day,
                )

                AttendanceRecord.objects.update_or_create(
                    student=student,
                    school_class=school_class,
                    attendance_date=attendance_date,
                    defaults={
                        "status": (
                            AttendanceStatus.PRESENT
                        ),
                        "marked_by": marked_by,
                    },
                )

                imported_count += 1

        return {
            "imported_count": imported_count,
            "skipped_count": skipped_count,
            "unmatched_students": unmatched_students,
        }